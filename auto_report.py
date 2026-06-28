#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
가톨릭혈액병원 경영팀 일일 현황 자동 생성 프로그램
매일 받는 zip 파일 -> 경영팀 표1(병원별 현황) + 표2(의사별 환자 현황) 자동 업데이트

사용법:
  python auto_report.py <zip파일경로> [출력경로]

  예시:
  python auto_report.py "매일 병원환자현황 리스트.zip"
  python auto_report.py "매일 병원환자현황 리스트.zip" "경영팀_20260625.xlsx"

필요 패키지:
  pip install openpyxl xlrd
"""

import zipfile, os, sys, shutil, re, tempfile
import openpyxl
from collections import defaultdict
from datetime import datetime

# ─── 고정 병상 수 (바뀌지 않는 값) ──────────────────────────────────────────
FIXED_BEDS = {
    'seoul':      {'dedicated': 244, 'transplant': 46, 'sterile': 74, 'general': 124},
}

# ─── 서울성모 병동 분류 ───────────────────────────────────────────────────────
SEOUL_GW_WARDS   = {'18층1병동', '18층2병동', '19층1병동', '19층2병동',
                    '20층1병동', '20층2병동', '21층1병동', '21층2병동'}
SEOUL_ICU_WARDS  = {'혈액계중환자실'}
SEOUL_ALL_WARDS  = SEOUL_GW_WARDS | SEOUL_ICU_WARDS
SEOUL_FULL_WARDS = SEOUL_ALL_WARDS | {'19층2병동', '21층1병동', '21층2병동'}

# 집계 대상 진료과 (서울성모 혈액병원 전문 진료과)
TARGET_DEPTS = {'혈액내과', '감염내과', '소아청소년과'}

# ─── 서울성모 전체 입원환자 집계용 의사 목록 (설명.xlsx 정의 기준) ──────────────
# 혈액내과 = 전체 (필터 없음)
INFEC_DOCS_SEOUL = {'이동건', '노덕희', '이래석', '조성연', '홍한터'}
PEDS_DOCS_SEOUL  = {'조빈', '정낙균', '이재욱', '유재원', '조수정'}
SEOUL_INPATIENT_WARDS = {
    '18층1병동', '18층2병동', '19층1병동', '19층2병동',
    '20층1병동', '20층2병동', '21층1병동', '21층2병동',
    '혈액계중환자실'
}

# ─── 기타 병원 병동명 ─────────────────────────────────────────────────────────
EUNPYEONG_WARDS = {'15층1병동', '16층2병동', '7A중환자실', '7B중환자실'}
YEOUIDO_WARDS   = {'11층1병동', '11층2병동', '12층1병동', '12층2병동',
                   '13층1병동', '13층2병동',
                   '신생아중환자실A', '신생아중환자실B', '신생아실'}
INCHEON_WARDS   = {'6층3병동', '7층3병동', '63병동', '73병동'}

# 기타 병원 이식병실 병동 (ICU와 신생아실 구분)
EUNPYEONG_TRANSPLANT = {'7A중환자실', '7B중환자실'}
YEOUIDO_TRANSPLANT   = set()   # 신생아중환자실은 혈액 이식병실 아님
INCHEON_TRANSPLANT   = set()

# ─── file2 식별용: 서울성모 병동별재원 병동 목록 ────────────────────────────
SEOUL_FILE2_WARDS = {
    '18층1병동', '18층2병동', '19층1병동', '19층2병동',
    '20층1병동', '20층2병동', '혈액계중환자실'
}

# ─── raw data 병동명 -> 경영팀 표2 컬럼명 매핑 ───────────────────────────────
WARD_TO_COL = {
    # 서울성모 GW
    '18층1병동': '18층1병동', '18층2병동': '18층2병동',
    '19층1병동': '19층1병동', '19층2병동': '19층2병동',
    '20층1병동': '20층1병동', '20층2병동': '20층2병동',
    '21층1병동': '21층1병동', '21층2병동': '21층2병동',
    # 서울성모 ICU
    '혈액계중환자실': '혈액계중환자실',
    '7A중환자실': '7층중환자실A', '7B중환자실': '7층중환자실B',
    # 소아
    '소아중환자실': '소아중환자실',
    # 기타 병원 GW
    '11층1병동': '11층1병동', '11층2병동': '11층2병동',
    '12층1병동': '12층1병동', '12층2병동': '12층2병동',
    '13층1병동': '13층1병동', '13층2병동': '13층2병동',
    '14층1병동': '14층1병동', '14층2병동': '14층2병동',
    '15층1병동': '15층1병동', '15층2병동': '15층2병동',
    '16층1병동': '16층1병동', '16층2병동': '16층2병동',
    '17층1병동': '17층1병동', '17층2병동': '17층2병동',
    '7층1병동':  '7층1병동',  '8층1병동':  '8층1병동',
    '8층2병동':  '8층2병동',  '10층1병동': '10층1병동',
    '10층2병동': '10층2병동',
}

# 경영팀 표2 ICU 컬럼 목록
ICU_COLS = {'혈액계중환자실', '7층중환자실A', '7층중환자실B',
            '내과중환자실', '내과중환자실B', '신경계중환자실',
            '심장계중환자실', '외과중환자실', '소아중환자실'}


# =============================================================================
# 1. RAW DATA 로드
# =============================================================================

def load_sheet(path):
    """xlsx 파일 로드 -> list of row tuples (header 제외)"""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        return rows[1:]
    except Exception as e:
        print("  [경고] 파일 로드 실패: {} -> {}".format(os.path.basename(path), e))
        return []


def load_xls_sheet(path):
    """구형 .xls 파일 로드 (xlrd 필요)"""
    try:
        import xlrd
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_index(0)
        return [ws.row_values(r) for r in range(ws.nrows)]
    except ImportError:
        print("  [경고] xlrd 미설치. pip install xlrd 실행 후 재시도")
        return []
    except Exception as e:
        print("  [경고] xls 로드 실패: {} -> {}".format(os.path.basename(path), e))
        return []


def _load_raw_with_header(path):
    """xlsx 파일을 (header_row, data_rows) 형태로 로드"""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], []
        return list(rows[0]), list(rows[1:])
    except Exception as e:
        print("  [경고] 파일 로드 실패: {} -> {}".format(os.path.basename(path), e))
        return [], []


def _identify_file_type(header_row, data_rows):
    """
    파일 헤더 + 데이터 내용만으로 파일 유형 식별 (파일명 절대 사용 안 함).

    식별 기준 (설명.xlsx 정의 기반):
      file4  : 헤더 col[1]에 '입원여부' 또는 col[2]에 '입원예약' 포함
      file3/7: 헤더 col[1] == 'BED'
               -> file7(응급실전일퇴원): 데이터 col[19]에 실제 날짜 존재
               -> file3(응급실재원): 그 외
      file5  : 헤더 col[1] == '환자번호'  (진단명 없음)
      file6  : 헤더 col[1] == '진단명' + col[21] 퇴원일자가 실제 날짜(!=9999)
      file2  : 헤더 col[1] == '진단명' + 병동 col[12] 값이 모두 서울병동 1개
      file1  : 나머지 (헤더 col[1] == '진단명')
      설명   : 위 어느 패턴도 해당 없음
    """
    if not header_row:
        return 'unknown'

    h = [str(c or '').strip().replace('\n', '') for c in header_row]
    h1 = h[1] if len(h) > 1 else ''
    h2 = h[2] if len(h) > 2 else ''

    # ── file4: 입원예약 ───────────────────────────────────────────────────────
    if '입원여부' in h1 or '입원예약' in h1 or '입원예약' in h2:
        return 'file4'

    # ── file3 / file7: ER 파일 (col[1]='BED') ────────────────────────────────
    if h1 == 'BED':
        # file7(ER discharge): col[6] status = circled-c (discharged)
        # file3(ER inpatient): col[6] is None or other char, never circled-c
        circ_c_bytes = 'ⓒ'  # ⓒ U+24D2 circled-c (file7 discharge status)
        has_circ_c = any(
            r[0] and len(r) > 6 and str(r[6] or '') == circ_c_bytes
            for r in data_rows
        )
        if has_circ_c:
            return 'file7'
        return 'file3'

    # ── file5: 전일입원 (col[1]='환자번호') ──────────────────────────────────
    if h1 == '환자번호':
        return 'file5'

    # ── file1 / file2 / file6 그룹: col[1]='진단명' ───────────────────────────
    if h1 == '진단명':
        # file6(전일퇴원): col[21] 퇴원일자가 실제 날짜 (재원은 9999-12-31)
        for r in data_rows[:20]:
            if r[0] and len(r) > 21 and r[21]:
                val = r[21]
                if hasattr(val, 'year') and 2020 < val.year < 9999:
                    return 'file6'

        # file2(병동별재원): 환자 병동이 서울성모 특정 병동 하나로만 구성
        wards = set()
        for r in data_rows[:50]:
            if r[0] and len(r) > 12 and r[12]:
                wards.add(str(r[12]).strip())
        if wards and wards.issubset(SEOUL_FILE2_WARDS):
            return 'file2'

        # 나머지: file1(과별재원)
        return 'file1'

    # 설명 파일 또는 알 수 없는 형식
    return '설명'


def _extract_report_date_from_data(data):
    """file6 퇴원일자 + 1일 = 보고 날짜 (파일명 의존 없음)"""
    from datetime import timedelta
    for r in data.get('file6', []):
        if r[0] and len(r) > 21 and r[21]:
            val = r[21]
            if hasattr(val, 'year') and 2020 < val.year < 9999:
                dis_date = val.date() if hasattr(val, 'date') else None
                if dis_date:
                    return datetime.combine(dis_date, datetime.min.time()) + timedelta(days=1)
    return datetime.today()


def load_all_data(zip_path):
    """zip 파일에서 모든 raw data 로드 — 파일명 아닌 내용으로 파일 유형 식별"""
    tmpdir = tempfile.mkdtemp()
    try:
        with zipfile.ZipFile(zip_path, 'r') as zf:
            zf.extractall(tmpdir)
        files = {f: os.path.join(tmpdir, f) for f in os.listdir(tmpdir)}
    except Exception as e:
        print("[오류] zip 파일 열기 실패: {}".format(e))
        sys.exit(1)

    data = {k: [] for k in ('file1', 'file2', 'file3', 'file4',
                             'file5', 'file6', 'file7', 'file8')}
    data['report_date'] = datetime.today()  # 임시; 아래에서 file6 기반으로 교체

    for fname, fpath in sorted(files.items()):
        fname_lower = fname.lower()

        # .xls (확장자 .xlsx 아닌 것) = file8 (BMT)
        if fname_lower.endswith('.xls') and not fname_lower.endswith('.xlsx'):
            rows = load_xls_sheet(fpath)
            data['file8'].extend(rows)  # header 포함; bmt 카운터에서 필터
            print("  [file8-BMT] {}".format(fname))
            continue

        if not fname_lower.endswith('.xlsx'):
            continue

        header, drows = _load_raw_with_header(fpath)
        ftype = _identify_file_type(header, drows)
        if ftype in data:
            data[ftype].extend(drows)
            print("  [{}] {} ({} 행)".format(ftype, fname[:45], len(drows)))
        else:
            print("  [건너뜀-{}] {}".format(ftype, fname[:45]))

    shutil.rmtree(tmpdir, ignore_errors=True)
    data = _finalize_data(data)
    data['report_date'] = _extract_report_date_from_data(data)
    return data


def load_files_from_dir(file_paths):
    """개별 xlsx/xls 파일 목록에서 데이터 로드 — 파일명 아닌 내용으로 파일 유형 식별"""
    data = {k: [] for k in ('file1', 'file2', 'file3', 'file4',
                             'file5', 'file6', 'file7', 'file8')}
    data['report_date'] = datetime.today()  # 임시; 아래에서 file6 기반으로 교체

    for fpath in sorted(file_paths):
        fname = os.path.basename(fpath)
        fname_lower = fname.lower()

        if fname_lower.endswith('.xls') and not fname_lower.endswith('.xlsx'):
            rows = load_xls_sheet(fpath)
            data['file8'].extend(rows)
            print("  [file8-BMT] {}".format(fname))
            continue

        if not fname_lower.endswith('.xlsx'):
            continue

        header, drows = _load_raw_with_header(fpath)
        ftype = _identify_file_type(header, drows)
        if ftype in data:
            data[ftype].extend(drows)
            print("  [{}] {} ({} 행)".format(ftype, fname[:45], len(drows)))
        else:
            print("  [건너뜀-{}] {}".format(ftype, fname[:45]))

    data = _finalize_data(data)
    data['report_date'] = _extract_report_date_from_data(data)
    return data


def _finalize_data(data):
    """중복 제거 + 통계 출력"""
    def dedup(rows, pid_col):
        seen = set(); result = []
        for r in rows:
            if r[0] and len(r) > pid_col and r[pid_col] not in seen:
                seen.add(r[pid_col])
                result.append(r)
        return result

    data['file1'] = dedup(data['file1'], 2)   # col[2]=환자번호
    data['file2'] = dedup(data['file2'], 2)
    data['file5'] = dedup(data['file5'], 1)   # file5 col[1]=환자번호
    data['file6'] = dedup(data['file6'], 2)

    print("  날짜: {}".format(data['report_date'].strftime('%Y-%m-%d')))
    print("  file1(과별재원): {}행 / file2(병동별): {}행".format(
        len(data['file1']), len(data['file2'])))
    print("  file3(응급재원): {}행 / file4(입원예약): {}행".format(
        len(data['file3']), len(data['file4'])))
    print("  file5(전일입원): {}행 / file6(전일퇴원): {}행 / file7(응급퇴원): {}행".format(
        len(data['file5']), len(data['file6']), len(data['file7'])))
    print("  file8(BMT): {}행".format(len(data['file8'])))
    return data


# =============================================================================
# 2. 경영팀 표1 계산
# =============================================================================

def room_suffix(room_str):
    """'20-207' -> 207"""
    if not room_str:
        return None
    m = re.search(r'-(\d+)$', str(room_str))
    return int(m.group(1)) if m else None


def is_ri(room_str):
    return 'ri' in str(room_str).lower() if room_str else False


def classify_patient(ward, room):
    """서울성모 환자를 병상 유형별로 분류. set 반환."""
    tags = set()
    if ward not in SEOUL_FULL_WARDS:
        return tags
    tags.add('total')
    rnum = room_suffix(room)

    if ward == '20층1병동':
        tags |= {'transplant', 'dedicated'}

    elif ward == '20층2병동':
        if rnum and 201 <= rnum <= 206:
            tags |= {'transplant', 'dedicated'}
        elif rnum and 207 <= rnum <= 214:
            tags |= {'sterile', 'dedicated'}
        else:
            tags.add('dedicated')

    elif ward == '19층2병동':
        tags |= {'sterile', 'dedicated'}

    elif ward == '19층1병동':
        ri = is_ri(room)
        suf = rnum % 100 if rnum else None
        if rnum and rnum in (115, 116):
            tags |= {'transplant', 'dedicated'}
        elif ri or (suf and suf in (8, 9, 10)):
            pass  # 전용 아님
        else:
            tags.add('dedicated')
        if not ri and suf and suf not in (8, 9, 10, 15, 16):
            tags.add('general_dedicated')

    elif ward in ('18층1병동', '18층2병동', '혈액계중환자실'):
        tags |= {'dedicated', 'general_dedicated'}

    elif ward in ('21층1병동', '21층2병동'):
        pass  # total 만 포함

    return tags


def calculate_table1(data):
    """경영팀 표1 값 계산 후 dict 반환"""
    from datetime import timedelta

    rd   = data['report_date']
    beds = FIXED_BEDS['seoul']

    def rnum(room_str):
        if not room_str: return None
        m = re.search(r'-?(\d+)$', str(room_str))
        return int(m.group(1)) if m else None

    def is_ri(room_str):
        return 'ri' in str(room_str).lower() if room_str else False

    def pct(a, b):
        return round(a / b * 100, 2) if b else 0

    # ─────────────────────────────────────────────────────────────────────────
    # 서울성모 병원
    # ─────────────────────────────────────────────────────────────────────────

    # ── 1. 전체 입원환자 수 (file1 과별조회, 설명.xlsx 정의 기준) ─────────────
    # 혈액내과(전체) + 감염내과(특정의사+서울병동) + 소아청소년과(특정의사)
    total = sum(1 for r in data['file1'] if r[0] and (
        str(r[10]).strip() == '혈액내과'
        or (str(r[10]).strip() == '감염내과'
            and str(r[11]).strip() in INFEC_DOCS_SEOUL
            and str(r[12]).strip() in SEOUL_INPATIENT_WARDS)
        or (str(r[10]).strip() == '소아청소년과'
            and str(r[11]).strip() in PEDS_DOCS_SEOUL)
    ))

    # ── 2. 이식/무균/일반 (file2 병동별조회, 병실번호 기준) ──────────────────
    # 이식병실: 201병동(전체) + 202병동(1-6호) + 19층1병동(115,116) — 설명.xlsx 정의 기준
    transplant = sum(1 for r in data['file2'] if r[0] and (
        str(r[12]).strip() == '20층1병동'
        or (str(r[12]).strip() == '20층2병동'
            and rnum(r[13]) is not None and 201 <= rnum(r[13]) <= 206)
        or (str(r[12]).strip() == '19층1병동'
            and rnum(r[13]) in (115, 116))
    ))

    # 무균병실: 192병동(전체) + 202병동(207-214호) — 설명.xlsx 정의 기준
    sterile = sum(1 for r in data['file2'] if r[0] and (
        str(r[12]).strip() == '19층2병동'
        or (str(r[12]).strip() == '20층2병동'
            and rnum(r[13]) is not None and 207 <= rnum(r[13]) <= 214)
    ))

    # 1)전용 일반병상: 181 + 182 + hicu + 191(8,9,10,15,16 제외) — 설명.xlsx 정의 기준
    gen_ded = sum(1 for r in data['file2'] if r[0] and (
        str(r[12]).strip() in ('18층1병동', '18층2병동', '혈액계중환자실')
        or (str(r[12]).strip() == '19층1병동'
            and rnum(r[13]) not in (108, 109, 110, 115, 116)
            and not is_ri(r[13]))
    ))

    # 전용병상 입원환자 수 — 설명.xlsx 정의: 201+202+181+182+hicu+191(108,109,110,ri 제외)
    # transplant+sterile+gen_ded 합산이 아닌 file2에서 직접 계산
    dedicated = sum(1 for r in data['file2'] if r[0] and (
        str(r[12]).strip() in ('20층1병동', '20층2병동', '18층1병동', '18층2병동', '혈액계중환자실')
        or (str(r[12]).strip() == '19층1병동'
            and rnum(r[13]) not in (108, 109, 110)
            and not is_ri(r[13]))
    ))
    general   = total - transplant - sterile
    gen_non   = general - gen_ded

    # ── 3. 응급실 (file3, 혈액+감염+소아 진료과) ─────────────────────────────
    er = sum(1 for r in data['file3']
             if r[0] and len(r) > 13 and str(r[13]).strip() in TARGET_DEPTS)

    # ── 4. 입원대기 (file4, 설명 기준) ──────────────────────────────────────
    # col[39]=예약진료과, col[15]=전문의
    # 혈액내과(전체) + 감염내과(특정의사) + 소아청소년과(특정의사)
    waiting = sum(1 for r in data['file4'] if r[0] and len(r) > 39 and (
        str(r[39]).strip() == '혈액내과'
        or (str(r[39]).strip() == '감염내과'
            and len(r) > 15 and str(r[15]).strip() in INFEC_DOCS_SEOUL)
        or (str(r[39]).strip() == '소아청소년과'
            and len(r) > 15 and str(r[15]).strip() in PEDS_DOCS_SEOUL)
    ))

    # ── 5. 전일입원 (file5, 서울병동 기준) ──────────────────────────────────
    prev_adm = sum(1 for r in data['file5']
                   if r[0] and len(r) > 11 and str(r[11]).strip() in SEOUL_INPATIENT_WARDS)

    # ── 6. 전일퇴원 (file6+file7, 설명 기준) ────────────────────────────────
    # file6: 혈액내과(전체) + 감염내과(INFEC_DOCS) + 소아청소년과(PEDS_DOCS)
    # file7: 혈액내과 중 내원일이 전일인 환자
    prev_date = (rd - timedelta(days=1)).date() if hasattr(rd, 'date') else None

    prev_dis = (
        sum(1 for r in data['file6']
            if r[0] and len(r) > 10 and str(r[10]).strip() == '혈액내과')
        + sum(1 for r in data['file6']
              if r[0] and len(r) > 11
              and str(r[10]).strip() == '감염내과'
              and str(r[11]).strip() in INFEC_DOCS_SEOUL)
        + sum(1 for r in data['file6']
              if r[0] and len(r) > 11
              and str(r[10]).strip() == '소아청소년과'
              and str(r[11]).strip() in PEDS_DOCS_SEOUL)
        + sum(1 for r in data['file7']
              if r[0] and len(r) > 19 and str(r[13]).strip() == '혈액내과'
              and r[19] and hasattr(r[19], 'date') and r[19].date() == prev_date)
    )

    # ── 7. BMT D-Day (file8, xls: col[10]=이름) ──────────────────────────────
    bmt = sum(1 for r in data['file8']
              if len(r) > 10 and r[10]
              and str(r[10]).strip() not in ('', '이름', '이름\n'))

    s = {
        'total':          total,
        'total_rate':     pct(total, beds['dedicated']),
        'dedicated':      dedicated,
        'dedicated_rate': pct(dedicated, beds['dedicated']),
        'transplant':     transplant,
        'transplant_rate':pct(transplant, beds['transplant']),
        'sterile':        sterile,
        'sterile_rate':   pct(sterile, beds['sterile']),
        'general':        general,
        'gen_ded':        gen_ded,
        'gen_non':        gen_non,
        'gen_rate':       pct(gen_ded, beds['general']),
        'er':             er,
        'total_er':       total + er,
        'waiting':        waiting,
        'prev_adm':       prev_adm,
        'prev_dis':       prev_dis,
        'prev_transfer':  0,
        'bmt':            bmt,
    }

    return {
        'date':  rd,
        'seoul': s,
    }


# =============================================================================
# 3. 경영팀 표2 계산
# =============================================================================

def calculate_disease_counts(data):
    """내과 입전의 담당질환별 환자 수 (표2 오른쪽 테이블)"""
    EXCLUDE_DOCS = {'임진조'}
    counts = {'AML': 0, 'ALL': 0, 'BMF/MPN': 0, 'MM': 0, 'Lymphoma': 0, 'Infection': 0, '중환자실': 0}

    for r in data['file1']:
        if not r[0]:
            continue
        dept = str(r[10]).strip()
        doc  = str(r[11]).strip() if r[11] else ''
        ward = str(r[12]).strip() if r[12] else ''
        diag = str(r[1]).strip()  if len(r) > 1 and r[1] else ''

        if doc in EXCLUDE_DOCS:
            continue
        if '기증자' in diag or '이식 상태' in diag:
            continue

        if dept == '혈액내과' and ward in SEOUL_INPATIENT_WARDS:
            d = diag
            if ward == '혈액계중환자실':
                counts['중환자실'] += 1
            elif (('골수성 백혈병' in d and '만성' not in d)
                  or '골수모구성' in d or '골수단구성' in d or '전골수구성' in d
                  or ('골수성 육종' in d and '림프' not in d)):
                counts['AML'] += 1
            elif ('림프모구성' in d or '림프모세포' in d or '버킷' in d
                  or ('전구물질' in d and '림프' in d)):
                counts['ALL'] += 1
            elif any(k in d for k in ['형성이상', '골수섬유증', '재생불량성', '저형성빈혈',
                                       '만성 골수성', '진성적혈구', '혈소판증가']):
                counts['BMF/MPN'] += 1
            elif any(k in d for k in ['골수종', '아밀로이드', '형질모세포']):
                counts['MM'] += 1
            else:
                # 나머지 혈액내과 (림프종, CLL, 조직구성, 기타 신생물 등)
                counts['Lymphoma'] += 1

        elif dept == '감염내과' and doc in INFEC_DOCS_SEOUL and ward in SEOUL_INPATIENT_WARDS:
            counts['Infection'] += 1

    return counts


def calculate_table2(data):
    """의사별·병동별 환자수 집계"""
    # 환자번호 -> 전문의 역인덱스 (응급환자 교차조회용)
    pid_to_doc = {}
    for row in data['file1']:
        if row[0] and row[2] and row[11]:
            pid_to_doc[row[2]] = row[11]

    doc_ward = defaultdict(lambda: defaultdict(int))

    # file1 병동별 집계
    for row in data['file1']:
        if not row[0]:
            continue
        doc, ward = row[11], row[12]
        if doc and ward:
            col = WARD_TO_COL.get(ward)
            if col:
                doc_ward[doc][col] += 1

    # file3 응급실 환자 -> 입원주치의(col32) 또는 file1 교차조회
    for row in data['file3']:
        if not row[0]:
            continue
        pid = row[7]
        doc = (row[32] if len(row) > 32 and row[32] else None) or pid_to_doc.get(pid)
        if doc:
            doc_ward[doc]['응급실재원'] += 1

    # GW / ICU / GW+ICU 합산
    result = {}
    for doc, wards in doc_ward.items():
        gw  = sum(v for k, v in wards.items() if k not in ICU_COLS and k != '응급실재원')
        icu = sum(v for k, v in wards.items() if k in ICU_COLS)
        result[doc] = dict(wards)
        result[doc]['GW+ICU'] = gw + icu
        result[doc]['GW']     = gw
        result[doc]['ICU']    = icu

    return result


# =============================================================================
# 4. EXCEL 업데이트
# =============================================================================

def find_template(script_dir):
    candidates = [
        os.path.join(script_dir, 'template.xlsx'),                                      # GitHub 배포용
        os.path.join(script_dir, 'template.xlsx.xlsx'),                                 # 확장자 중복 업로드된 경우
        os.path.join(script_dir, '교수님 자료', 'AI_agent_개발_우선순위 포함.xlsx'),    # 로컬용
        os.path.join(script_dir, 'AI_agent_개발_우선순위 포함.xlsx'),
    ]
    for c in candidates:
        if os.path.exists(c):
            return c
    for f in os.listdir(script_dir):
        if '우선순위' in f and f.endswith('.xlsx'):
            return os.path.join(script_dir, f)
    return None


def update_table1(ws, t1):
    """경영팀 표1 시트 업데이트"""
    # 날짜 셀 업데이트
    for row in ws.iter_rows(min_row=1, max_row=5):
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.value = t1['date']

    s = t1['seoul']

    EXACT_LABELS = {
        '전체 입원환자 수':            s['total'],
        '전용대비 입원환자 가동율(%)':  s['total_rate'],
        '전용병상 입원환자 수':         s['dedicated'],
        '전용병상 가동율(%)':           s['dedicated_rate'],
        '전용 이식병실 가동율(%)':      s['transplant_rate'],
        '전용 무균병실 가동율(%)':      s['sterile_rate'],
        '일반병실 입원환자 수':         s['general'],
        '1) 전용병상 입원환자  수':     s['gen_ded'],
        '2) 비전용병상 입원환자 수':    s['gen_non'],
        '전용일반병실 가동율(%)':       s['gen_rate'],
        '입원+응급실':                  s['total_er'],
        '입원대기 (외래)':              s['waiting'],
        '전일입원*':                    s['prev_adm'],
        '전일퇴원*':                    s['prev_dis'],
        '전일전원*':                    s['prev_transfer'],
        'BMT D-Day':                    s['bmt'],
    }
    PARTIAL_LABELS = {
        '전용대비 입원환자 가동율':  s['total_rate'],
        '전용병상 가동율':           s['dedicated_rate'],
        '전용 이식병실 가동율':      s['transplant_rate'],
        '전용 무균병실 가동율':      s['sterile_rate'],
        '전용일반병실 가동율':       s['gen_rate'],
    }

    section = None

    for row in ws.iter_rows(min_row=1):
        b    = str(row[1].value or '').strip() if len(row) > 1 else ''
        rnum = row[0].row
        a    = str(row[0].value or '').strip() if row[0].value else ''

        # 섹션 감지
        if '이식병실' in a:   section = 'transplant'
        elif '무균병실' in a: section = 'sterile'
        elif '일반병실' in a: section = 'general'
        elif a == '응급실':   section = 'er'
        elif a in ('전체 입원', '기타현황'): section = None

        # 섹션별 입원환자 수
        if b == '입원환자 수':
            if section == 'transplant':
                ws.cell(rnum, 3).value = s['transplant']
            elif section == 'sterile':
                ws.cell(rnum, 3).value = s['sterile']
            continue

        # 응급실 행
        if a == '응급실' and b == '응급실':
            ws.cell(rnum, 3).value = s['er']
            continue

        # label 매칭 → 서울성모 컬럼(3)만 기록
        val = EXACT_LABELS.get(b) if b in EXACT_LABELS else PARTIAL_LABELS.get(b)
        if val is not None:
            ws.cell(rnum, 3).value = val


def update_table2(ws, t2):
    """경영팀 표2 시트 업데이트"""
    # 실제 집계 대상 컬럼만 col_idx에 포함 (옆 테이블 컬럼 제외)
    VALID_COL_NAMES = {'GW+ICU', 'GW', 'ICU', '응급실재원'} | set(WARD_TO_COL.values()) | ICU_COLS

    col_idx = {}
    header_row_num = 2
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        if row[1] and '주치의' in str(row[1]):
            header_row_num = i
            for j, v in enumerate(row):
                nm = str(v).strip() if v else ''
                if nm and (nm in ('진료과', '주치의') or nm in VALID_COL_NAMES):
                    col_idx[nm] = j + 1
            break

    if not col_idx:
        print("  [경고] 표2 헤더를 찾지 못했습니다.")
        return

    doc_col = col_idx.get('주치의', 2)
    updated = 0
    _all_sums_cache = {}

    for row in ws.iter_rows(min_row=header_row_num + 1):
        cell_val = row[doc_col - 1].value
        if not cell_val:
            continue
        doc = str(cell_val).strip()
        if doc not in t2:
            continue

        counts = t2[doc]
        updated += 1

        for col_name, col_num in col_idx.items():
            if col_name in ('GW+ICU', 'GW', 'ICU'):
                ws.cell(row[0].row, col_num).value = counts.get(col_name, 0)
            elif col_name == '응급실재원':
                ws.cell(row[0].row, col_num).value = counts.get('응급실재원', 0)
            elif col_name in VALID_COL_NAMES:
                ws.cell(row[0].row, col_num).value = counts.get(col_name, 0)

    print("  표2: {}명 의사 업데이트 완료".format(updated))

    # ── 합계 행 직접 계산 (SUM 공식 의존 제거) ──────────────────────────
    dept_col = col_idx.get('진료과', 1)
    sum_cols = [cn for cn in col_idx if cn not in ('진료과', '주치의')]

    # 1) 일반 의사 행에서 과별 합계 누적
    dept_sums = defaultdict(lambda: defaultdict(int))
    for row in ws.iter_rows(min_row=header_row_num + 1):
        dept_cell = row[dept_col - 1].value
        doc_cell  = row[doc_col - 1].value
        if not dept_cell or not doc_cell:
            continue
        dept = str(dept_cell).strip()
        doc  = str(doc_cell).strip()
        if doc not in t2:
            continue
        counts = t2[doc]
        for cn in sum_cols:
            dept_sums[dept][cn] += counts.get(cn, 0)

    # 2) '합계' 행 찾아서 기록 (총합계 포함)
    for row in ws.iter_rows(min_row=header_row_num + 1):
        label = ''
        for ci in range(min(3, len(row))):
            v = row[ci].value
            if v and '합계' in str(v):
                label = str(v).strip()
                break
        if not label:
            continue

        if '총합계' in label or '총 합계' in label:
            # 전체 합계: 모든 과 dept_sums 합산
            all_sums = defaultdict(int)
            for ds in dept_sums.values():
                for cn, val in ds.items():
                    all_sums[cn] += val
            for cn, col_num in col_idx.items():
                if cn in sum_cols:
                    ws.cell(row[0].row, col_num).value = all_sums.get(cn, 0)

            # ── 병동별 병상가동율% 는 아래 별도 루프에서 처리 ─────────────
            _all_sums_cache.update(all_sums)
        else:
            dept_name = label.replace(' 합계', '').strip()
            if dept_name not in dept_sums:
                continue
            sums = dept_sums[dept_name]
            for cn, col_num in col_idx.items():
                if cn in sum_cols:
                    ws.cell(row[0].row, col_num).value = sums.get(cn, 0)

    # ── 병동별 병상가동율% 행: '병동별' 텍스트 스캔해서 직접 값 쓰기 ──────
    WARD_BEDS = {
        '18층1병동': 44, '18층2병동': 44, '19층1병동': 39,
        '19층2병동': 44, '20층1병동': 32, '20층2병동': 36,
        '혈액계중환자실': 5,
    }
    all_sums = _all_sums_cache
    for row in ws.iter_rows(min_row=header_row_num + 1, max_row=header_row_num + 40):
        a = str(row[0].value or '').strip()
        if '병동별' in a and '병상가동율' in a:
            for ward, beds in WARD_BEDS.items():
                col_num = col_idx.get(ward)
                if col_num and beds:
                    rate = round(all_sums.get(ward, 0) / beds * 100, 2)
                    ws.cell(row[0].row, col_num).value = rate
            print("  병동별 병상가동율 업데이트 완료 (행 {})".format(row[0].row))
            break


def fix_column_widths(ws):
    """
    숫자 열: 내용 길이 + 1, 최소 4, 최대 8 (### 방지 + 뚱뚱함 방지)
    텍스트 열: 건드리지 않음
    """
    from openpyxl.utils import get_column_letter
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_num_len = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            if isinstance(cell.value, float):
                s = "{:.1f}".format(cell.value)
            elif isinstance(cell.value, int):
                s = str(cell.value)
            else:
                continue
            if len(s) > max_num_len:
                max_num_len = len(s)
        if max_num_len > 0:
            ws.column_dimensions[col_letter].width = min(max(max_num_len + 1, 4), 8)


def _delete_other_hospital_cols(ws):
    for col_idx in (6, 5, 4):
        ws.delete_cols(col_idx)


def update_disease_table(ws, disease_counts):
    DISEASE_KEYS = {"AML", "ALL", "BMF/MPN", "MM", "Lymphoma", "Infection", "중환자실"}
    for row in ws.iter_rows(min_row=1, max_row=50):
        if len(row) < 43:
            continue
        lbl = str(row[42].value or "").strip()
        if lbl in DISEASE_KEYS:
            ws.cell(row[42].row, 44).value = disease_counts.get(lbl, 0)
    print("  disease table: {}".format(disease_counts))


def create_clean_excel(t1, t2, template_path, output_path, data=None):
    import shutil
    shutil.copy2(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    keep = {name for name in wb.sheetnames if "푔01" in name or "푔02" in name
            or "푔01" in name or "푔02" in name}
    keep2 = {name for name in wb.sheetnames if "표1" in name or "표2" in name}
    keep = keep | keep2
    # fallback: keep sheets with digit 1 or 2 in name that look like report sheets
    if not keep:
        keep = set(wb.sheetnames[:2])
    for name in list(wb.sheetnames):
        if name not in keep:
            del wb[name]
    for name in list(wb.sheetnames):
        if "표1" in name or ("1" in name and "표" in name):
            ws = wb[name]
            update_table1(ws, t1)
            _delete_other_hospital_cols(ws)
            break
    for name in list(wb.sheetnames):
        if "표2" in name or ("2" in name and "표" in name):
            ws = wb[name]
            update_table2(ws, t2)
            if data is not None:
                disease_counts = calculate_disease_counts(data)
                update_disease_table(ws, disease_counts)
            break
    wb.save(output_path)


def write_output(template_path, t1, t2, output_path):
    import shutil
    shutil.copy2(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    names = wb.sheetnames
    for name in names:
        if "표1" in name or ("1" in name and "표" in name):
            update_table1(wb[name], t1)
            break
    for name in names:
        if "표2" in name or ("2" in name and "표" in name):
            update_table2(wb[name], t2)
            break
    wb.save(output_path)
    print("done: {}".format(output_path))


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    if len(sys.argv) < 2:
        zips = [f for f in os.listdir(script_dir) if f.endswith(".zip")]
        if not zips:
            print("usage: python auto_report.py <zip> [output]")
            sys.exit(1)
        zip_path = os.path.join(script_dir, zips[0])
    else:
        zip_path = sys.argv[1]
        if not os.path.isabs(zip_path):
            zip_path = os.path.join(script_dir, zip_path)
    template_path = find_template(script_dir)
    if not template_path:
        print("[ERROR] template not found")
        sys.exit(1)
    output_path = sys.argv[2] if len(sys.argv) >= 3 else None
    if output_path and not os.path.isabs(output_path):
        output_path = os.path.join(script_dir, output_path)
    data = load_all_data(zip_path)
    if output_path is None:
        ds = data["report_date"].strftime("%Y%m%d")
        output_path = os.path.join(script_dir, "report_{}.xlsx".format(ds))
    t1 = calculate_table1(data)
    t2 = calculate_table2(data)
    create_clean_excel(t1, t2, template_path, output_path, data=data)
    print("Done:", output_path)


if __name__ == '__main__':
    main()
